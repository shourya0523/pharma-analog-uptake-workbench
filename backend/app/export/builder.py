from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.db.models import (
    DatapointORM,
    DrugJobORM,
    DrugProfileFieldORM,
    ExportORM,
    QualityCheckORM,
    SourceDocumentORM,
    UnresolvedQuarterORM,
    ValidationTaskORM,
)
from app.domain.models import new_id
from app.storage.filestore import FileStore


QUARTERLY_HEADERS = [
    "drug_name",
    "period",
    "fiscal_year",
    "fiscal_quarter",
    "calendar_year",
    "calendar_quarter",
    "value_reported",
    "value_normalized_usd_millions",
    "currency",
    "unit",
    "metric",
    "revenue_scope",
    "geography",
    "formulation",
    "route_of_administration",
    "source_url",
    "source_quote",
    "source_support",
    "extraction_method",
    "confidence_score",
    "validation_status",
    "reviewer_notes",
    "issue_flags",
]


class ExportBuilder:
    def __init__(self, db: Session, file_store: FileStore) -> None:
        self.db = db
        self.file_store = file_store

    async def export_product_workbook(self, job_id: str) -> ExportORM:
        job = self.db.get(DrugJobORM, job_id)
        if not job:
            raise ValueError("job not found")
        wb = Workbook()

        # Sheet 1 quarterly
        ws = wb.active
        ws.title = "Quarterly Revenue"
        ws.append(QUARTERLY_HEADERS)
        for d in job.datapoints:
            ws.append(
                [
                    job.drug_name,
                    d.period,
                    d.fiscal_year,
                    d.fiscal_quarter,
                    d.calendar_year,
                    d.calendar_quarter,
                    d.value_reported,
                    d.value_normalized_usd_millions,
                    d.currency,
                    d.unit,
                    d.metric,
                    d.revenue_scope,
                    d.geography,
                    d.formulation,
                    d.route_of_administration,
                    d.source_url,
                    d.source_quote,
                    d.source_support,
                    d.extraction_method,
                    d.confidence_score,
                    d.validation_status,
                    d.reviewer_notes,
                    ",".join(d.issue_flags or []),
                ]
            )

        ws2 = wb.create_sheet("Source Audit Log")
        ws2.append(
            [
                "source_id",
                "drug_name",
                "source_type",
                "source_title",
                "source_url",
                "source_date",
                "filing_type",
                "accession_number",
                "page_or_section",
                "retrieval_status",
                "parsing_status",
                "relevant_datapoints_found",
                "notes",
            ]
        )
        for s in job.sources:
            ws2.append(
                [
                    s.id,
                    job.drug_name,
                    s.source_type,
                    s.source_title,
                    s.source_url,
                    s.source_date,
                    s.filing_type,
                    s.accession_number,
                    s.page_or_section,
                    s.retrieval_status,
                    s.parsing_status,
                    s.relevant_datapoints_found,
                    s.notes,
                ]
            )

        ws3 = wb.create_sheet("Unresolved Quarter Tracker")
        ws3.append(
            [
                "drug_name",
                "period",
                "reason_unresolved",
                "sources_checked",
                "recommended_next_step",
                "confidence_that_unavailable",
                "reviewer_notes",
            ]
        )
        for u in job.unresolved_quarters:
            ws3.append(
                [
                    job.drug_name,
                    u.period,
                    u.reason_unresolved,
                    ",".join(u.sources_checked or []),
                    u.recommended_next_step,
                    u.confidence_that_unavailable,
                    u.reviewer_notes,
                ]
            )

        ws4 = wb.create_sheet("Drug Profile")
        ws4.append(["field", "value", "source_url", "confidence", "validation_status"])
        for f in job.profile_fields:
            cit = f.citation_json or {}
            ws4.append([f.field, f.value, cit.get("source_url"), cit.get("confidence"), f.validation_status])

        ws5 = wb.create_sheet("Quality Checks")
        ws5.append(["issue_type", "severity", "affected_datapoint", "explanation", "recommended_action", "status"])
        for q in job.quality_checks:
            ws5.append([q.issue_type, q.severity, q.affected_datapoint, q.explanation, q.recommended_action, q.status])

        buf = io.BytesIO()
        wb.save(buf)
        key = f"exports/{job.run_id}/{job.id}/product_workbook.xlsx"
        await self.file_store.put(key, buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        exp = ExportORM(id=new_id(), run_id=job.run_id, job_id=job.id, format="product_workbook", storage_key=key)
        self.db.add(exp)
        self.db.commit()
        return exp

    async def export_powerbi_csvs(self, run_id: str) -> list[ExportORM]:
        jobs = self.db.query(DrugJobORM).filter_by(run_id=run_id).all()
        exports: list[ExportORM] = []

        def write_csv(name: str, headers: list[str], rows: list[list]) -> ExportORM:
            out = io.StringIO()
            w = csv.writer(out)
            w.writerow(headers)
            w.writerows(rows)
            key = f"exports/{run_id}/powerbi/{name}"
            # sync put via asyncio loop caller
            return key, out.getvalue().encode()

        # products
        prod_rows = []
        for j in jobs:
            fields = {f.field: f.value for f in j.profile_fields}
            prod_rows.append(
                [
                    j.id,
                    j.drug_name,
                    j.generic_name,
                    j.manufacturer,
                    fields.get("therapeutic_area"),
                    fields.get("moa"),
                    fields.get("roa"),
                    fields.get("fda_approval_date"),
                    j.completeness_pct,
                    j.status,
                ]
            )
        key, data = write_csv(
            "products.csv",
            [
                "job_id",
                "product_name",
                "generic_name",
                "manufacturer",
                "therapeutic_area",
                "moa",
                "roa",
                "fda_approval_date",
                "completeness_pct",
                "validation_status",
            ],
            prod_rows,
        )
        await self.file_store.put(key, data, "text/csv")
        exp = ExportORM(id=new_id(), run_id=run_id, format="products_csv", storage_key=key)
        self.db.add(exp)
        exports.append(exp)

        q_rows = []
        for j in jobs:
            for d in j.datapoints:
                q_rows.append(
                    [
                        j.drug_name,
                        d.period,
                        d.value_normalized_usd_millions,
                        d.source_url,
                        d.source_quote,
                        d.confidence_score,
                        d.validation_status,
                        d.revenue_scope,
                    ]
                )
        key, data = write_csv(
            "quarterly_revenue.csv",
            [
                "drug_name",
                "period",
                "value_normalized_usd_millions",
                "source_url",
                "source_quote",
                "confidence_score",
                "validation_status",
                "revenue_scope",
            ],
            q_rows,
        )
        await self.file_store.put(key, data, "text/csv")
        exp = ExportORM(id=new_id(), run_id=run_id, format="quarterly_revenue_csv", storage_key=key)
        self.db.add(exp)
        exports.append(exp)

        self.db.commit()
        return exports


class TemplateMapper:
    """Stub for future official Excel workbook mapping."""

    def infer(self, workbook_bytes: bytes) -> dict:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(max_row=1, values_only=True))
            headers = list(rows[0]) if rows else []
            sheets[name] = {"headers": headers}
        return {"sheets": sheets, "compatible": True, "unmapped_required": []}
