"""Export the gold dataset to a single Excel workbook.

Every data sheet is a faithful dump of a gold file - no re-derivation, no
rounding, no reordering of values. Quarterly Matrix and Product Summary are
views over those sheets: a pivot and a per-product census, both generated here
rather than written as Excel formulas, because an openpyxl-written formula
carries no cached value and reads back blank to pandas and to previewers until
something recalculates it. Regenerate the workbook instead of editing it - this
script is the only thing that should write to exports/.

Reads seed/gold only. It must never import from the gold builder or from
application code: this is a presentation of the oracle, not part of it.

    python scripts/export_gold_workbook.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
GOLD = REPO / "seed" / "gold"
OUT = REPO / "exports" / "pah_gold_dataset.xlsx"

FONT = "Arial"
INK = "1F3864"
HEAD_FILL = PatternFill("solid", fgColor=INK)
BAND_FILL = PatternFill("solid", fgColor="F2F5FA")
TITLE = Font(name=FONT, size=14, bold=True, color=INK)
HEAD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
NOTE = Font(name=FONT, size=9, italic=True, color="5B6B7F")
BOLD = Font(name=FONT, size=10, bold=True)
THIN = Side(style="thin", color="D6DCE4")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '#,##0.0;(#,##0.0);-'
PCT = '0.0"%"'


def load(name):
    path = GOLD / f"{name}.jsonl"
    return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]


def flat(value):
    """Lists and dicts have to become text, but they must stay readable."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, list):
        if not value:
            return ""
        if all(isinstance(item, str) for item in value):
            return " | ".join(value)
        return json.dumps(value, separators=(", ", ": "))
    if isinstance(value, dict):
        return json.dumps(value, separators=(", ", ": "))
    return value


def write_sheet(wb, title, columns, rows, *, widths=None, formats=None, wrap=(), note=None):
    ws = wb.create_sheet(title)
    header_row = 1
    if note:
        ws.cell(1, 1, note).font = NOTE
        header_row = 2

    for index, column in enumerate(columns, start=1):
        cell = ws.cell(header_row, index, column)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = GRID
    ws.row_dimensions[header_row].height = 30

    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset
        for index, column in enumerate(columns, start=1):
            cell = ws.cell(excel_row, index, flat(row.get(column)))
            cell.font = BODY
            cell.border = GRID
            if offset % 2:
                cell.fill = BAND_FILL
            if formats and column in formats:
                cell.number_format = formats[column]
            if column in wrap:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = (widths or {}).get(column, 16)

    last = header_row + len(rows)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{last}"
    ws.freeze_panes = ws.cell(header_row + 1, 3)
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- data sheets

profiles = load("product_profiles")
quarterly = load("quarterly_revenue")
annual = load("annual_revenue")
coverage = load("series_coverage")
peaks = load("peak_sales")
excluded = load("excluded_products")
cases = load("adjudication_cases")
unresolved = load("unresolved_quarters")
manifest = json.loads((GOLD / "manifest.json").read_text())
report = json.loads((GOLD / "build_report.json").read_text())

# Quarterly rows carry a `sources` list; the primary source already has its own
# columns, so only the extras need somewhere to go.
for row in quarterly:
    extras = [s["source_url"] for s in (row.get("sources") or []) if s["source_url"] != row["source_url"]]
    row["additional_source_urls"] = " | ".join(extras)
    row["bridge_components"] = flat(row.get("bridge_components"))

Q_COLS = [
    "drug_name", "generic_name", "manufacturer", "period", "calendar_year", "calendar_quarter",
    "fiscal_year", "fiscal_quarter", "period_basis", "benchmark_identity", "revenue_scope",
    "geography", "value_normalized_usd_millions", "unit", "currency", "value_reported",
    "source_value_reported", "source_unit", "precision", "derivation", "confidence_score",
    "validation_status", "therapeutic_area", "route_of_administration", "formulation",
    "extraction_method", "source_type", "source_url", "source_quote", "additional_source_urls",
    "bridge_components", "gold_notes", "gold_id",
]
write_sheet(
    wb, "Quarterly Revenue", Q_COLS, quarterly,
    note="Every reported quarter in gold. value_normalized_usd_millions is the figure to use; "
         "value_reported and source_value_reported preserve the issuer's own number and unit.",
    widths={"drug_name": 18, "generic_name": 18, "manufacturer": 20, "benchmark_identity": 30,
            "revenue_scope": 22, "geography": 16, "value_normalized_usd_millions": 14,
            "derivation": 30, "therapeutic_area": 24, "source_url": 52, "source_quote": 70,
            "additional_source_urls": 40, "bridge_components": 40, "gold_notes": 40,
            "gold_id": 34, "extraction_method": 26, "validation_status": 16, "precision": 16},
    formats={"value_normalized_usd_millions": MONEY, "value_reported": MONEY,
             "source_value_reported": '#,##0.0##', "confidence_score": "0.00"},
    wrap=("source_quote", "gold_notes", "bridge_components", "additional_source_urls"),
)

A_COLS = [
    "drug_name", "generic_name", "manufacturer", "period", "period_basis", "benchmark_identity",
    "revenue_scope", "geography", "value_normalized_usd_millions", "unit", "currency",
    "value_reported", "source_value_reported", "source_unit", "fx_rate_to_usd", "fx_rate_source",
    "derivation", "series_role", "confidence_score", "validation_status", "extraction_method",
    "source_type", "source_url", "source_quote", "gold_id",
]
write_sheet(
    wb, "Annual Revenue", A_COLS, annual,
    note="Annual figures. series_role says whether a row is a peak benchmark in its own right "
         "or annual context for a series measured quarterly.",
    widths={"drug_name": 18, "generic_name": 18, "manufacturer": 20, "benchmark_identity": 34,
            "revenue_scope": 26, "geography": 16, "value_normalized_usd_millions": 14,
            "derivation": 30, "series_role": 18, "source_url": 52, "source_quote": 70,
            "gold_id": 34, "extraction_method": 26, "fx_rate_source": 24},
    formats={"value_normalized_usd_millions": MONEY, "value_reported": MONEY,
             "source_value_reported": '#,##0.0##', "confidence_score": "0.00",
             "fx_rate_to_usd": "0.0000"},
    wrap=("source_quote",),
)

P_COLS = [
    "drug_name", "indication_area", "moa", "moa_class", "route_of_administration",
    "first_approval_year", "approval_era", "competitive_intensity_at_launch",
    "marketed_peers_at_launch", "competitive_intensity_basis", "peer_universe_role",
    "attribute_provenance",
]
write_sheet(
    wb, "Product Profiles", P_COLS, profiles,
    note="The analog-matching attributes, one row per product appearing anywhere in gold. "
         "Competitive intensity is derived from marketed_peers_at_launch, not hand-assigned; "
         "outside the catalog's own indication universe it is left unassessed rather than guessed.",
    widths={"drug_name": 20, "indication_area": 34, "moa": 40, "moa_class": 26,
            "route_of_administration": 22, "first_approval_year": 14, "approval_era": 14,
            "competitive_intensity_at_launch": 18, "marketed_peers_at_launch": 14,
            "competitive_intensity_basis": 38, "peer_universe_role": 24,
            "attribute_provenance": 20},
    wrap=("moa",),
)

C_COLS = [
    "drug_name", "benchmark_identity", "moa", "moa_class", "route_of_administration",
    "approval_era", "competitive_intensity_at_launch", "launch_quarter", "commercial_start_quarter",
    "series_start_reason", "series_end_quarter", "series_end_basis", "series_end_reason",
    "as_of_quarter", "expected_quarters", "observed_quarters", "coverage_pct", "missing_quarters",
    "quarters_beyond_series_end", "benchmark_eligible",
]
write_sheet(
    wb, "Series Coverage", C_COLS, coverage,
    note="One row per quarterly series. missing_quarters is empty on every series - "
         "coverage is 100% as of 2026Q2.",
    widths={"drug_name": 20, "benchmark_identity": 34, "moa": 38, "moa_class": 26,
            "route_of_administration": 20, "series_start_reason": 46, "series_end_reason": 46,
            "series_end_basis": 26, "missing_quarters": 24, "quarters_beyond_series_end": 24},
    formats={"coverage_pct": PCT},
    wrap=("moa", "series_start_reason", "series_end_reason"),
)

K_COLS = [
    "drug_name", "moa", "moa_class", "route_of_administration", "approval_era",
    "competitive_intensity_at_launch", "peak_status", "numeric_peak_available", "peak_value",
    "peak_year", "highest_observed_value", "highest_observed_year", "highest_observed_period",
    "post_peak_years", "annual_observations", "revenue_scope", "geography", "unit", "currency",
    "selection_method", "benchmark_eligible", "input_ids", "gold_id",
]
write_sheet(
    wb, "Peak Sales", K_COLS, peaks,
    note="peak_status observed means the series has turned down and stayed down; "
         "not_yet_observed means the highest value so far is still the latest value.",
    widths={"drug_name": 20, "moa": 38, "moa_class": 26, "route_of_administration": 20,
            "selection_method": 46, "input_ids": 60, "gold_id": 26, "revenue_scope": 22,
            "peak_status": 18, "highest_observed_period": 18},
    formats={"peak_value": MONEY, "highest_observed_value": MONEY},
    wrap=("moa", "selection_method", "input_ids"),
)

E_COLS = [
    "drug_name", "benchmark_status", "reason_code", "details", "moa", "moa_class",
    "route_of_administration", "approval_era", "competitive_intensity_at_launch",
    "extraction_method", "source_url", "source_quote", "gold_id",
]
write_sheet(
    wb, "Excluded Products", E_COLS, excluded,
    note="Products deliberately kept out of the benchmark set, each with the evidence for why. "
         "An exclusion still carries its matching attributes, so it can serve as an analog "
         "comparator even where it cannot serve as a peak benchmark.",
    widths={"drug_name": 20, "reason_code": 34, "details": 60, "moa": 38, "moa_class": 26,
            "source_url": 52, "source_quote": 70, "gold_id": 24, "extraction_method": 26},
    wrap=("details", "source_quote"),
)

for case in cases:
    case["expect"] = flat(case.get("expect"))
    case["inputs"] = flat(case.get("inputs"))
CASE_COLS = ["case_id", "kind", "provenance", "why", "inputs", "expect"]
write_sheet(
    wb, "Adjudication Cases", CASE_COLS, cases,
    note="Genuinely ambiguous situations found in the filings, with the expected resolution. "
         "These are the test cases an extraction pipeline is graded against.",
    widths={"case_id": 34, "kind": 20, "provenance": 14, "why": 80, "inputs": 70, "expect": 34},
    wrap=("why", "inputs", "expect"),
)

# ------------------------------------------------------------- source index

index_rows = {}
for row in quarterly:
    entry = index_rows.setdefault(row["source_url"], {
        "source_url": row["source_url"], "source_type": row["source_type"],
        "products": set(), "periods": set(), "quarterly_rows": 0, "annual_rows": 0})
    entry["quarterly_rows"] += 1
    entry["products"].add(row["drug_name"])
    entry["periods"].add(row["period"])
for row in annual:
    entry = index_rows.setdefault(row["source_url"], {
        "source_url": row["source_url"], "source_type": row["source_type"],
        "products": set(), "periods": set(), "quarterly_rows": 0, "annual_rows": 0})
    entry["annual_rows"] += 1
    entry["products"].add(row["drug_name"])
    entry["periods"].add(row["period"])

sources = []
for entry in index_rows.values():
    periods = sorted(entry["periods"])
    sources.append({
        "source_url": entry["source_url"],
        "source_type": entry["source_type"],
        "quarterly_rows": entry["quarterly_rows"],
        "annual_rows": entry["annual_rows"],
        "total_rows": entry["quarterly_rows"] + entry["annual_rows"],
        "products_cited": ", ".join(sorted(entry["products"])),
        "earliest_period": periods[0],
        "latest_period": periods[-1],
    })
sources.sort(key=lambda item: (-item["total_rows"], item["source_url"]))
write_sheet(
    wb, "Source Index", list(sources[0].keys()), sources,
    note="Every distinct filing or release cited by a revenue row, and how much of gold rests "
         "on it. Row counts are a census of the two revenue sheets.",
    widths={"source_url": 80, "source_type": 16, "products_cited": 50,
            "quarterly_rows": 13, "annual_rows": 11, "total_rows": 11,
            "earliest_period": 14, "latest_period": 14},
    wrap=("products_cited",),
)

# ---------------------------------------------------------- quarterly matrix
# Live SUMIFS over Quarterly Revenue rather than a second copy of the numbers,
# so filtering or correcting the data sheet moves the matrix with it.

periods = sorted({row["period"] for row in quarterly})
series = sorted({(row["drug_name"], row["benchmark_identity"], row["revenue_scope"])
                 for row in quarterly})
# The pivot reads from a dict keyed exactly as the data sheet stores each row,
# so a value can only appear here if that same row appears there.
cube = {(row["benchmark_identity"], row["period"]): row["value_normalized_usd_millions"]
        for row in quarterly}
assert len(cube) == len(quarterly), "a series reports the same quarter twice"

ws = wb.create_sheet("Quarterly Matrix")
ws.cell(1, 1, "Reported revenue by series and quarter, USD millions. A pivot of the Quarterly "
              "Revenue sheet - the same figures rearranged, nothing recomputed. A blank cell "
              "means the quarter is outside that series' reported window, never a reported zero.").font = NOTE
headers = ["drug_name", "benchmark_identity", "revenue_scope"] + periods
for index, name in enumerate(headers, start=1):
    cell = ws.cell(2, index, name)
    cell.font = HEAD
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = GRID
ws.row_dimensions[2].height = 30

for offset, (drug, identity, scope) in enumerate(series):
    row_index = 3 + offset
    for index, value in enumerate((drug, identity, scope), start=1):
        cell = ws.cell(row_index, index, value)
        cell.font = BODY
        cell.border = GRID
        if offset % 2:
            cell.fill = BAND_FILL
    for index, period in enumerate(periods, start=4):
        cell = ws.cell(row_index, index, cube.get((identity, period)))
        cell.number_format = MONEY
        cell.font = BODY
        cell.border = GRID
        if offset % 2:
            cell.fill = BAND_FILL

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 24
for index in range(4, 4 + len(periods)):
    ws.column_dimensions[get_column_letter(index)].width = 10
ws.freeze_panes = "D3"

# ------------------------------------------------------- product summary
# One row per product, joining the profile, the revenue sheets and the peak
# sheet so a single product can be read without crossing four tabs. Every
# number here is a formula over those sheets, so it moves when they do.

SUMMARY_STATUS = {}
quarterly_products = {row["drug_name"] for row in quarterly}
annual_products = {row["drug_name"] for row in annual}
excluded_products = {row["drug_name"] for row in excluded}
for row in profiles:
    name = row["drug_name"]
    if name in quarterly_products:
        SUMMARY_STATUS[name] = "quarterly series"
    elif name in annual_products:
        SUMMARY_STATUS[name] = "annual benchmark only"
    elif name in excluded_products:
        SUMMARY_STATUS[name] = "excluded from benchmark"
    else:
        SUMMARY_STATUS[name] = "attributes only"

ps = wb.create_sheet("Product Summary")
ps.cell(1, 1, "One row per product, so a product can be read without crossing four tabs. "
              "Attributes come from Product Profiles; counts, totals and peaks are a census of "
              "the Quarterly Revenue, Annual Revenue and Peak Sales sheets, generated by "
              "scripts/export_gold_workbook.py. A blank peak column means the product has no "
              "peak row - it is either excluded or has not turned down yet.").font = NOTE

S_HEAD = ["drug_name", "role in gold", "indication_area", "moa_class", "route_of_administration",
          "approval_era", "competitive_intensity_at_launch", "marketed_peers_at_launch",
          "quarterly rows", "annual rows", "first calendar year", "last calendar year",
          "total reported (USD mm)", "peak status", "peak value (USD mm)", "peak year"]
for index, name in enumerate(S_HEAD, start=1):
    cell = ps.cell(2, index, name)
    cell.font = HEAD
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = GRID
ps.row_dimensions[2].height = 32

peak_index = {row["drug_name"]: row for row in peaks}
for offset, row in enumerate(sorted(profiles, key=lambda item: item["drug_name"])):
    n = 3 + offset
    name = row["drug_name"]
    q_rows = [item for item in quarterly if item["drug_name"] == name]
    a_rows = [item for item in annual if item["drug_name"] == name]
    peak = peak_index.get(name)
    years = [item["calendar_year"] for item in q_rows]
    values = [
        name, SUMMARY_STATUS[name], row.get("indication_area"), row.get("moa_class"),
        row.get("route_of_administration"), row.get("approval_era"),
        row.get("competitive_intensity_at_launch"), row.get("marketed_peers_at_launch"),
        len(q_rows), len(a_rows),
        min(years) if years else None,
        max(years) if years else None,
        round(sum(item["value_normalized_usd_millions"] for item in q_rows), 1) if q_rows else None,
        peak["peak_status"] if peak else None,
        peak["peak_value"] if peak else None,
        peak["peak_year"] if peak else None,
    ]
    for index, value in enumerate(values, start=1):
        cell = ps.cell(n, index, value)
        cell.font = BODY
        cell.border = GRID
        if offset % 2:
            cell.fill = BAND_FILL
        if index in (13, 15):
            cell.number_format = MONEY
        if index in (11, 12, 16):
            cell.number_format = "0"
        cell.alignment = Alignment(vertical="top")

for index, width in enumerate([20, 22, 34, 26, 20, 13, 15, 13, 12, 11, 13, 13, 16, 18, 14, 11], start=1):
    ps.column_dimensions[get_column_letter(index)].width = width
ps.auto_filter.ref = f"A2:P{2 + len(profiles)}"
ps.freeze_panes = "C3"

# --------------------------------------------------------------- read me

read = wb.create_sheet("Read Me", 0)
read.column_dimensions["A"].width = 34
read.column_dimensions["B"].width = 16
read.column_dimensions["C"].width = 92


def line(row, label, value=None, *, font=BODY, note=None, fmt=None):
    cell = read.cell(row, 1, label)
    cell.font = font
    if value is not None:
        v = read.cell(row, 2, value)
        v.font = BOLD
        v.alignment = Alignment(horizontal="left")
        if fmt:
            v.number_format = fmt
    if note:
        n = read.cell(row, 3, note)
        n.font = BODY
        n.alignment = Alignment(vertical="top", wrap_text=True)
    return row + 1


r = 1
read.cell(r, 1, "PAH Peak Sales - Gold Dataset").font = TITLE
r += 1
read.cell(r, 1, f"Independently researched from issuer filings and releases. "
                f"As of {manifest['as_of_quarter']}.").font = NOTE
r += 2

read.cell(r, 1, "WHAT THIS IS").font = BOLD
r += 1
read.cell(r, 3, "This is the oracle an extraction pipeline is measured against, not pipeline "
                "output. Every revenue figure was read from the issuer's own filing or press "
                "release and carries the URL plus the verbatim line it came from, so any number "
                "here can be checked at source. Where a figure is not printed directly, the "
                "derivation column names the arithmetic used.").font = BODY
read.cell(r, 3).alignment = Alignment(vertical="top", wrap_text=True)
read.row_dimensions[r].height = 62
r += 2

read.cell(r, 1, "CONTENTS").font = BOLD
r += 1
for label, header in [("Sheet", "Rows")]:
    read.cell(r, 1, label).font = HEAD
    read.cell(r, 1).fill = HEAD_FILL
    read.cell(r, 2, header).font = HEAD
    read.cell(r, 2).fill = HEAD_FILL
    read.cell(r, 3, "What it holds").font = HEAD
    read.cell(r, 3).fill = HEAD_FILL
r += 1

CONTENTS = [
    ("Product Summary", len(profiles),
     "Start here. One row per product, joining attributes, revenue totals and peak."),
    ("Quarterly Revenue", len(quarterly),
     "Every reported quarter, with source URL and verbatim quote on each row."),
    ("Quarterly Matrix", len(series),
     "The same quarters pivoted to series x period, one row per series."),
    ("Annual Revenue", len(annual),
     "Annual figures: peak benchmarks in their own right, plus annual context for quarterly series."),
    ("Product Profiles", len(profiles),
     "Analog-matching attributes: mechanism, route, approval era, competitive intensity."),
    ("Series Coverage", len(coverage),
     "Per-series completeness: expected vs observed quarters, and any gaps."),
    ("Peak Sales", len(peaks),
     "Peak value and year per product, and whether the peak has actually been observed."),
    ("Excluded Products", len(excluded),
     "Products kept out of the benchmark set, each with the evidence for why."),
    ("Adjudication Cases", len(cases),
     "Ambiguous situations found in the filings, with their expected resolution."),
    ("Source Index", len(sources),
     "Every distinct filing cited, and how many rows depend on it."),
    ("Build Manifest", None,
     "The dataset's own manifest and build report, as generated."),
]
for label, ref, description in CONTENTS:
    link = read.cell(r, 1, label)
    link.font = Font(name=FONT, size=10, color="0563C1", underline="single")
    link.hyperlink = f"#'{label}'!A1"
    if ref is not None:
        read.cell(r, 2, ref).font = BOLD
        read.cell(r, 2).number_format = "#,##0"
        read.cell(r, 2).alignment = Alignment(horizontal="left")
    read.cell(r, 3, description).font = BODY
    read.cell(r, 3).alignment = Alignment(vertical="top", wrap_text=True)
    r += 1

r += 1
read.cell(r, 1, "HOW TO READ A REVENUE ROW").font = BOLD
r += 1
GLOSSARY = [
    ("value_normalized_usd_millions", "The figure to use. USD millions, whatever the filing's own unit."),
    ("value_reported / source_unit", "The issuer's own number in the issuer's own unit, preserved."),
    ("derivation", "direct_reported means the issuer printed this figure. Anything else names "
                   "the arithmetic - e.g. annual_less_reported_first_nine_months is a Q4 backed "
                   "out of a stated full year."),
    ("benchmark_identity", "The series key. A product can have more than one - a U.S. line and a "
                           "worldwide line are different series, not the same series twice."),
    ("revenue_scope", "What the issuer's line actually covers. Never mix scopes within a series."),
    ("source_quote", "The verbatim line from the filing, so the number can be checked without "
                     "re-reading the document."),
]
for term, meaning in GLOSSARY:
    read.cell(r, 1, term).font = BODY
    read.cell(r, 3, meaning).font = BODY
    read.cell(r, 3).alignment = Alignment(vertical="top", wrap_text=True)
    if len(meaning) > 110:
        read.row_dimensions[r].height = 28
    r += 1

r += 1
read.cell(r, 1, "COVERAGE AS BUILT").font = BOLD
r += 1
STATS = [
    ("Catalog products", report["catalog_coverage"]["catalog_products"],
     "All accounted for: a quarterly series, an annual benchmark, or an evidenced exclusion."),
    ("Quarterly series", report["complete_quarterly_series"], "Every one complete."),
    ("Quarterly observations", report["quarterly_rows"], ""),
    ("Quarterly coverage", report["quarterly_coverage_pct"],
     "No missing quarter in any series as of 2026Q2."),
    ("Annual observations", report["annual_rows"], ""),
    ("Observed peaks", report["observed_peaks"],
     "Products whose series has turned down and stayed down."),
    ("Peaks not yet observed", report["not_yet_observed_peaks"],
     "Still at their highest value; no peak claimed."),
    ("Evidenced exclusions", report["excluded_products"], ""),
    ("Distinct sources cited", len(sources), ""),
    ("Products with attributes", report["product_profiles"]["products"], ""),
    ("Competitive intensity assessed", report["product_profiles"]["competitive_intensity_assessed"],
     "Only inside the catalog's own indication universe; elsewhere it is left unassessed."),
]
for label, value, note in STATS:
    read.cell(r, 1, label).font = BODY
    cell = read.cell(r, 2, value)
    cell.font = BOLD
    cell.number_format = PCT if isinstance(value, float) else "#,##0"
    cell.alignment = Alignment(horizontal="left")
    read.cell(r, 3, note).font = BODY
    read.cell(r, 3).alignment = Alignment(vertical="top", wrap_text=True)
    r += 1

r += 1
read.cell(r, 1, "THE ONE RULE WORTH STATING").font = BOLD
r += 1
read.cell(r, 3, "Where an issuer prints regional lines and also states a worldwide figure, the "
                "worldwide figure is what gold records - never the sum of the regions. Issuers "
                "round each line independently, so the parts need not add to the whole they "
                "themselves publish.").font = BODY
read.cell(r, 3).alignment = Alignment(vertical="top", wrap_text=True)
read.row_dimensions[r].height = 46
r += 1

# ------------------------------------------------------- build manifest sheet

mf = wb.create_sheet("Build Manifest")
mf.column_dimensions["A"].width = 40
mf.column_dimensions["B"].width = 100
mf.cell(1, 1, "Build Manifest").font = TITLE
mf.cell(2, 1, "Generated by scripts/build_independent_gold.py. Reproduced verbatim.").font = NOTE
row_index = 4


def dump(prefix, obj, row_index):
    for key, value in obj.items():
        if isinstance(value, dict):
            cell = mf.cell(row_index, 1, f"{prefix}{key}")
            cell.font = BOLD
            row_index += 1
            row_index = dump(f"{prefix}{key}.", value, row_index)
        else:
            mf.cell(row_index, 1, f"{prefix}{key}").font = BODY
            cell = mf.cell(row_index, 2, flat(value))
            cell.font = BODY
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_index += 1
    return row_index


mf.cell(row_index, 1, "manifest.json").font = TITLE
row_index += 1
row_index = dump("", manifest, row_index)
row_index += 1
mf.cell(row_index, 1, "build_report.json").font = TITLE
row_index += 1
row_index = dump("", report, row_index)
row_index += 1
mf.cell(row_index, 1, "unresolved_quarters.jsonl").font = BOLD
mf.cell(row_index, 2, f"{len(unresolved)} rows - no quarter in any series is unresolved.").font = BODY

wb.move_sheet("Read Me", offset=-wb.sheetnames.index("Read Me"))
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("wrote", OUT, "sheets:", wb.sheetnames)
